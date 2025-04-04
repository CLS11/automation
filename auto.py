import openpyxl as xl
from openpyxl.chart import BarChart, Reference

#Function for automating multiple workbooks
def processWorkbook(filename):
    workbook = xl.load_workbook(filename)
    sheet = workbook['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        correctedPriceCell = sheet.cell(row, 4)
        correctedPriceCell.value = cell.value * 0.9

    values = Reference(
        sheet,  
        min_row = 2, 
        max_row = sheet.max_row,
        min_col = 4,
        max_col = 4)

    chart = BarChart()
    chart.add_data(values) 
    sheet.add_chart(chart, 'e2')

    workbook.save(filename)



